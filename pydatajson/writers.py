#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""Módulo 'writers' de pydatajson

Contiene los métodos para escribir
- diccionarios con metadatos de catálogos a formato JSON, así como
- listas de diccionarios ("tablas") en formato CSV o XLSX
"""

from __future__ import print_function, unicode_literals, with_statement

import io
import json
import logging
import os

import openpyxl as pyxl
import unicodecsv as csv
from openpyxl.styles import Font
from openpyxl.utils import column_index_from_string
from six import string_types, text_type, moves, iteritems

from . import helpers

logger = logging.getLogger('pydatajson')


def write_tables(tables, path, column_styles=None, cell_styles=None,
                 tables_fields=None, tables_names=None):
    """ Exporta un reporte con varias tablas en CSV o XLSX.

    Si la extensión es ".csv" se crean varias tablas agregando el nombre de la
    tabla al final del "path". Si la extensión es ".xlsx" todas las tablas se
    escriben en el mismo excel.

    Args:
        table (dict of (list of dicts)): Conjunto de tablas a ser exportadas
            donde {
                "table_name": [{
                    "field_name1": "field_value1",
                    "field_name2": "field_value2",
                    "field_name3": "field_value3"
                }]
            }
        path (str): Path al archivo CSV o XLSX de exportación.
    """
    assert isinstance(path, string_types), "`path` debe ser un string"
    assert isinstance(tables, dict), "`table` es dict de listas de dicts"

    # Deduzco el formato de archivo de `path` y redirijo según corresponda.
    suffix = path.split(".")[-1]
    if suffix == "csv":
        for table_name, table in tables:
            root_path = "".join(path.split(".")[:-1])
            table_path = "{}_{}.csv".format(root_path, table_name)
            _write_csv_table(table, table_path)

    elif suffix == "xlsx":
        return _write_xlsx_table(tables, path, column_styles, cell_styles,
                                 tables_fields=tables_fields,
                                 tables_names=tables_names)

    else:
        raise ValueError("""
{} no es un sufijo reconocido. Pruebe con .csv o.xlsx""".format(suffix))


def write_table(table, path, column_styles=None, cell_styles=None):
    """ Exporta una tabla en el formato deseado (CSV o XLSX).

    La extensión del archivo debe ser ".csv" o ".xlsx", y en función de
    ella se decidirá qué método usar para escribirlo.

    Args:
        table (list of dicts): Tabla a ser exportada.
        path (str): Path al archivo CSV o XLSX de exportación.
    """
    assert isinstance(path, string_types), "`path` debe ser un string"
    assert isinstance(table, list), "`table` debe ser una lista de dicts"

    # si la tabla está vacía, no escribe nada
    if len(table) == 0:
        logger.warning("Tabla vacia: no se genera ninguna archivo.")
        return

    # Sólo sabe escribir listas de diccionarios con información tabular
    if not helpers.is_list_of_matching_dicts(table):
        raise ValueError("""
La lista ingresada no esta formada por diccionarios con las mismas claves.""")

    # Deduzco el formato de archivo de `path` y redirijo según corresponda.
    suffix = path.split(".")[-1]
    if suffix == "csv":
        return _write_csv_table(table, path)
    elif suffix == "xlsx":
        return _write_xlsx_table(table, path, column_styles, cell_styles)
    else:
        raise ValueError("""
{} no es un sufijo reconocido. Pruebe con .csv o.xlsx""".format(suffix))


def _write_csv_table(table, path):
    if len(table) == 0:
        logger.warning("No se puede crear un CSV con una tabla vacía.")
        return

    headers = table[0].keys()

    with open(path, 'wb') as target_file:
        writer = csv.DictWriter(csvfile=target_file, fieldnames=headers,
                                lineterminator="\n", encoding='utf-8')
        writer.writeheader()
        for row in table:
            writer.writerow(row)


def _apply_styles_to_ws(ws, column_styles=None, cell_styles=None):
    # dict de las columnas que corresponden a cada campo
    header_row = next(ws.rows)
    headers_cols = {cell.value: cell.column for cell in header_row}

    # aplica estilos de columnas
    if column_styles:
        for col, properties in iteritems(column_styles):
            # la col puede ser "A" o "nombre_campo"
            col = headers_cols.get(col, col)
            for prop_name, prop_value in iteritems(properties):
                setattr(ws.column_dimensions[col], prop_name, prop_value)

    # aplica estilos de celdas
    if cell_styles:
        for i in moves.xrange(1, ws.max_row + 1):
            for j in moves.xrange(1, ws.max_column + 1):
                cell = ws.cell(row=i, column=j)

                # si el valor es una URL válida, la celda es un hyperlink
                if helpers.validate_url(cell.value):
                    cell.hyperlink = cell.value
                    cell.font = Font(underline='single', color='0563C1')

                for cell_style in cell_styles:
                    match_all = (
                        "col" not in cell_style and
                        "row" not in cell_style
                    )
                    match_row = (
                        "row" in cell_style and
                        cell_style["row"] == i
                    )
                    match_col = (
                        "col" in cell_style and
                        column_index_from_string(
                            headers_cols.get(cell_style["col"],
                                             cell_style["col"])) == j
                    )
                    if match_all or match_row or match_col:
                        for prop_name, prop_value in iteritems(cell_style):
                            if prop_name != "col" and prop_name != "row":
                                setattr(cell, prop_name, prop_value)


def _write_xlsx_table(tables, path, column_styles=None, cell_styles=None,
                      tables_fields=None, tables_names=None):
    column_styles = column_styles or {}
    cell_styles = cell_styles or {}
    wb = pyxl.Workbook()

    if isinstance(tables, dict):
        ws_names = []

        # primero se usa `tables_names`, y después las extra que pueda haber
        if tables_names:
            ws_names.extend(tables_names)
            for key in tables.keys():
                if key not in ws_names:
                    ws_names.append(key)

        # se agregan los nombres de las tablas que falten
        else:
            ws_names = tables.keys()

        wb.remove(wb.active)

        for table_name in ws_names:
            table = tables.get(table_name)
            column_styles_sheet = column_styles.get(table_name)
            cell_styles_sheet = cell_styles.get(table_name)

            _list_table_to_ws(
                wb, table, table_name, column_styles_sheet,
                cell_styles_sheet,
                fields=tables_fields.get(table_name) if tables_fields else None
            )

    else:
        _list_table_to_ws(wb, tables, column_styles=column_styles,
                          cell_styles=cell_styles)

    wb.save(path)


def _list_table_to_ws(wb, table, table_name=None, column_styles=None,
                      cell_styles=None, fields=None):
    if len(table) == 0 and not fields:
        logger.warning("No se puede crear una hoja Excel con una tabla vacía.")
        return
    elif len(table) == 0 and fields:
        # la primer fila de la tabla está vacía
        table.append({field: None for field in fields})

    if table_name:
        ws = wb.create_sheet(title=table_name)
    else:
        ws = wb.active

    headers = []
    # primero se usan los fields pasados, y después los extra que pueda haber
    if fields:
        headers.extend(fields)
        for key in table[0].keys():
            if key not in headers:
                headers.append(key)
    # se usan los headers de la primera fila para toda la tabla
    else:
        headers = list(table[0].keys())

    ws.append(headers)

    for index, row in enumerate(table):
        row_values = []
        for header in headers:
            # si el header no está en la fila, tiene valor nulo
            value = row.get(header)
            if isinstance(value, list):
                row_values.append(",".join(value))
            else:
                row_values.append(value)

        ws.append(row_values)

    _apply_styles_to_ws(ws, column_styles, cell_styles)


def write_json(obj, path):
    """Escribo un objeto a un archivo JSON con codificación UTF-8."""
    obj_str = text_type(json.dumps(obj, indent=4, separators=(",", ": "),
                                   ensure_ascii=False))

    helpers.ensure_dir_exists(os.path.dirname(path))

    with io.open(path, "w", encoding='utf-8') as target:
        target.write(obj_str)


def write_json_catalog(catalog, path):
    """Escribe el catálogo en JSON.

    Args:
        catalog (DataJson): Catálogo de datos.
        path (str): Directorio absoluto donde se crea el archivo XLSX.
    """
    write_json(catalog, path)


XLSX_FIELDS = {
    "catalog": [
        "catalog_identifier",
        "catalog_title",
        "catalog_description",
        "catalog_publisher_name",
        "catalog_publisher_mbox",
        "catalog_issued",
        "catalog_modified",
        "catalog_language",
        "catalog_superThemeTaxonomy",
        "catalog_license",
        "catalog_homepage",
        "catalog_rights",
        "catalog_spatial"
    ],
    "dataset": [
        "dataset_identifier",
        "dataset_title",
        "dataset_description",
        "dataset_publisher_name",
        "dataset_publisher_mbox",
        "dataset_contactPoint_fn",
        "dataset_contactPoint_hasEmail",
        "dataset_superTheme",
        "dataset_theme",
        "dataset_keyword",
        "dataset_accrualPeriodicity",
        "dataset_issued",
        "dataset_modified",
        "dataset_language",
        "dataset_spatial",
        "dataset_temporal",
        "dataset_landingPage",
        "dataset_license",
        "dataset_source",
        "dataset_accessLevel"
    ],
    "distribution": [
        "dataset_identifier",
        "dataset_title",
        "distribution_identifier",
        "distribution_title",
        "distribution_description",
        "distribution_downloadURL",
        "distribution_fileName",
        "distribution_format",
        "distribution_accessURL",
        "distribution_mediaType",
        "distribution_license",
        "distribution_byteSize",
        "distribution_issued",
        "distribution_modified",
        "distribution_rights"
    ],
    "field": [
        "dataset_identifier",
        "dataset_title",
        "distribution_identifier",
        "distribution_title",
        "field_title",
        "field_type",
        "field_description"
    ],
    "theme": [
        "theme_id",
        "theme_label",
        "theme_description"
    ]
}


def _tabulate_nested_dict(nested_dict_row, field_root="dataset",
                          parents_roots=[]):
    table_dict_row = {}

    for key, value in nested_dict_row.items():
        if not isinstance(value, dict):

            has_root = False
            for root in parents_roots + [field_root]:
                if key.startswith(root):
                    has_root = True

            if has_root:
                table_dict_row[key] = value
            else:
                table_dict_row["{}_{}".format(field_root, key)] = value

        else:
            tabulated_keys = _tabulate_nested_dict(value, field_root=key)
            for nested_key, nested_value in tabulated_keys.items():
                if nested_key.startswith(field_root):
                    table_dict_row[nested_key] = value
                else:
                    table_dict_row["{}_{}".format(
                        field_root, nested_key)] = nested_value

    return table_dict_row


def _generate_dataset_table(catalog):
    headers = []
    datasets = []

    # tabula diccionarios con estructura, como listas planas de diccionarios
    for dataset in catalog.get_datasets(exclude_meta_fields=["distribution"]):
        tab_dataset = _tabulate_nested_dict(dataset, "dataset")
        datasets.append(tab_dataset)

        # agrega todas las keys nuevas que no estén trackeadas
        for key in tab_dataset:
            if key not in headers:
                headers.append(key)

    # agrega "nones" para todos aquellos datasets que no tengan alguna key
    for dataset in datasets:
        for header in headers:
            if header not in dataset:
                dataset[header] = None

    return datasets


def _generate_distribution_table(catalog):
    headers = []
    distributions = []

    # tabula diccionarios con estructura, como listas planas de diccionarios
    for distribution in catalog.get_distributions(
            exclude_meta_fields=["field"]):
        tab_distribution = _tabulate_nested_dict(
            distribution, "distribution", ["dataset"])
        tab_distribution["dataset_title"] = catalog.get_dataset(
            tab_distribution["dataset_identifier"]).get("title")
        distributions.append(tab_distribution)

        # agrega todas las keys nuevas que no estén trackeadas
        for key in tab_distribution:
            if key not in headers:
                headers.append(key)

    # agrega "nones" para todos aquellos datasets que no tengan alguna key
    for distribution in distributions:
        for header in headers:
            if header not in distribution:
                distribution[header] = None

    return distributions


def _generate_field_table(catalog):
    headers = []
    fields = []

    # tabula diccionarios con estructura, como listas planas de diccionarios
    for field in catalog.get_fields():
        tab_field = _tabulate_nested_dict(
            field, "field", ["dataset", "distribution"])
        tab_field["dataset_title"] = catalog.get_dataset(
            tab_field["dataset_identifier"]).get("title")
        tab_field["distribution_title"] = catalog.get_distribution(
            tab_field["distribution_identifier"]).get("title")
        fields.append(tab_field)

        # agrega todas las keys nuevas que no estén trackeadas
        for key in tab_field:
            if key not in headers:
                headers.append(key)

    # agrega "nones" para todos aquellos datasets que no tengan alguna key
    for field in fields:
        for header in headers:
            if header not in field:
                field[header] = None

    return fields


def _generate_theme_table(catalog):
    headers = []
    themes = []
    catalog_themes = catalog.get_themes() or []
    # tabula diccionarios con estructura, como listas planas de diccionarios
    for theme in catalog_themes:
        tab_theme = _tabulate_nested_dict(theme, "theme")
        themes.append(tab_theme)

        # agrega todas las keys nuevas que no estén trackeadas
        for key in tab_theme:
            if key not in headers:
                headers.append(key)

    # agrega "nones" para todos aquellos datasets que no tengan alguna key
    for theme in themes:
        for header in headers:
            if header not in theme:
                theme[header] = None

    return themes


def write_xlsx_catalog(catalog, path, xlsx_fields=None):
    """Escribe el catálogo en Excel.

    Args:
        catalog (DataJson): Catálogo de datos.
        path (str): Directorio absoluto donde se crea el archivo XLSX.
        xlsx_fields (dict): Orden en que los campos del perfil de metadatos
            se escriben en cada hoja del Excel.
    """

    xlsx_fields = xlsx_fields or XLSX_FIELDS
    catalog_dict = {}

    catalog_dict["catalog"] = [
        _tabulate_nested_dict(catalog.get_catalog_metadata(
            exclude_meta_fields=["themeTaxonomy"]),
            "catalog")
    ]
    catalog_dict["dataset"] = _generate_dataset_table(catalog)
    catalog_dict["distribution"] = _generate_distribution_table(catalog)
    catalog_dict["field"] = _generate_field_table(catalog)
    catalog_dict["theme"] = _generate_theme_table(catalog)

    write_tables(
        catalog_dict, path, tables_fields=xlsx_fields,
        tables_names=["catalog", "dataset", "distribution", "field", "theme"]
    )
